#!/usr/bin/env python3
"""
视频历史数据导入脚本
将抖音后台导出的Excel数据导入 video_performance_record 表

用法:
    # 生成SQL文件（推荐，不需要数据库连接）
    python3 scripts/import_video_performance.py --mode sql --input <excel文件路径>

    # 直连MySQL（需安装 pymysql: pip3 install pymysql）
    python3 scripts/import_video_performance.py --mode db \
        --input <excel文件路径> \
        --host 127.0.0.1 --port 3306 \
        --user root --password xxx --db marketing_person

输出:
    --mode sql: scripts/output/video_performance_insert.sql
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl。执行: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# SKU标注规则（按优先级匹配）
# ============================================================
SKU_RULES = [
    ("种籽气垫",  r"种籽气垫|种子气垫"),
    ("野藤气垫",  r"野藤气垫|野藤"),
    ("900目粉霜", r"900目|粉霜"),
    ("腮红系列",  r"腮红"),
    ("口红系列",  r"口红|唇"),
    ("其他",      r""),  # 兜底
]

# ============================================================
# 钩子类型标注规则（按优先级匹配）
# 注意：技术卖点型放在价格机制型之前，避免"抢先购"误判
# ============================================================
HOOK_RULES = [
    ("明星背书型",  r"枫稳|林枫松|陈稳|明星同款|代言|侯明昊|徐海乔|海陆|陈瑶"),
    ("技术卖点型",  r"\d+小时持妆|\d+小时控油|微米级|换皮级|遮瑕力|上脸无痕|扒脸|控油耐汗|防水防汗|四效合一|养肤底妆|熟龄肌|法令纹|轻薄如肤"),
    ("场景植入型",  r"旅游|早八|约会|通勤|长白山|雪屋|户外|出门|上班|出行|运动|偶遇|探店"),
    ("情绪共鸣型",  r"谁懂|服气|含金量|震惊|巨巨|真的好|太好看|太好用|惊艳|没想到|居然|竟然"),
    ("价格机制型",  r"买\d发\d|拍\d发\d|拍一发|买一发|\d个\d+多|拍1享|买1享|闭眼入|划算|免费试用|抽黄金|抽奖|拍1发"),
    ("活动通知型",  r"新品|上市|限定|奔赴|倒计时|预售|上新|首发|首次|痛楼|线下|直播"),
    ("种草测评型",  r"种草|测评|真实|原相机|真实上脸|实测|好用推荐|必买|必入|闭眼入"),
    ("其他",        r""),  # 兜底
]

# ============================================================
# 辅助函数
# ============================================================

def parse_num(v, default=0.0):
    """解析可能含逗号或%的数字"""
    if v is None:
        return default
    s = str(v).replace(',', '').replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return default


def parse_duration_sec(v):
    """'2秒' → 2, '1分30秒' → 90, '0秒' → 0"""
    if v is None:
        return 0
    s = str(v).strip()
    mins = re.search(r'(\d+)分', s)
    secs = re.search(r'(\d+)秒', s)
    m = int(mins.group(1)) if mins else 0
    sec = int(secs.group(1)) if secs else 0
    return m * 60 + sec


def parse_datetime(v):
    """解析发布时间"""
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    s = str(v).strip()
    for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return None


def annotate_sku(title: str) -> str:
    if not title:
        return "其他"
    for label, pattern in SKU_RULES:
        if pattern and re.search(pattern, title):
            return label
    return "其他"


def annotate_hook(title: str) -> tuple[str, str]:
    """返回 (hook_type, content_pattern)"""
    if not title:
        return "其他", "其他"
    matched = []
    for label, pattern in HOOK_RULES:
        if pattern and re.search(pattern, title):
            matched.append(label)
    hook_type = matched[0] if matched else "其他"
    content_pattern = "+".join(matched) if matched else "其他"
    return hook_type, content_pattern


def compute_composite_score(row: dict, max_vals: dict) -> float:
    """
    多目标归一化评分：
      引流直播GMV      × 40%
      引流直播次数     × 25%
      完播率           × 15%
      平均观看秒数     × 10%
      看后搜GMV        × 10%
    """
    def norm(key, weight):
        m = max_vals.get(key, 1) or 1
        return (row.get(key, 0) / m) * weight

    score = (
        norm('live_gmv', 0.40) +
        norm('live_traffic', 0.25) +
        norm('completion_rate', 0.15) +
        norm('avg_watch_sec', 0.10) +
        norm('post_search_gmv', 0.10)
    )
    return round(score, 6)


def escape_sql(v) -> str:
    if v is None:
        return "NULL"
    s = str(v).replace("\\", "\\\\").replace("'", "\\'").replace("\n", " ").replace("\r", "")
    return f"'{s}'"


# ============================================================
# 主逻辑
# ============================================================

def load_excel(path: str) -> list[dict]:
    print(f"读取文件: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    print(f"  共 {len(rows)} 条记录")
    return rows


def transform(rows: list[dict]) -> list[dict]:
    """Excel行 → 标准化字段dict"""
    records = []
    for r in rows:
        title = str(r.get('视频标题') or '')
        hook_type, content_pattern = annotate_hook(title)
        sku_tag = annotate_sku(title)

        rec = {
            'video_id':         str(r.get('视频ID') or ''),
            'account_id':       str(r.get('达人抖音号') or ''),
            'account_name':     str(r.get('达人昵称') or ''),
            'title':            title[:990],
            'publish_time':     parse_datetime(r.get('发布时间')),
            'video_url':        str(r.get('播放链接') or '')[:1020],
            'is_promoted':      1 if str(r.get('是否投放') or '') == '是' else 0,
            'views':            int(parse_num(r.get('视频观看次数'))),
            'likes':            int(parse_num(r.get('点赞数'))),
            'comments':         int(parse_num(r.get('评论数'))),
            'saves':            int(parse_num(r.get('收藏数'))),
            'shares':           int(parse_num(r.get('转发数'))),
            'follows':          int(parse_num(r.get('点击关注次数'))),
            'completion_rate':  parse_num(r.get('完播率')),
            'avg_watch_sec':    parse_duration_sec(r.get('平均观看时长')),
            'live_gmv':         parse_num(r.get('引流直播间用户支付金额(元)')),
            'live_exposure':    int(parse_num(r.get('直播间入口曝光次数'))),
            'live_traffic':     int(parse_num(r.get('引流直播间次数'))),
            'live_orders':      int(parse_num(r.get('引流直播成交订单数'))),
            'live_cpm_gmv':     parse_num(r.get('入口千次曝光用户支付金额(元)')),
            'post_search_gmv':  parse_num(r.get('看后搜用户支付金额(元)')),
            'post_search_exp':  int(parse_num(r.get('看后搜商品曝光次数'))),
            'post_search_click':int(parse_num(r.get('看后搜商品点击次数'))),
            'post_search_orders':int(parse_num(r.get('看后搜成交订单数'))),
            'direct_gmv':       parse_num(r.get('用户支付金额(元)')),
            'direct_orders':    int(parse_num(r.get('成交订单数'))),
            'refund_amount':    parse_num(r.get('退款金额(元)')),
            'store_orders':     int(parse_num(r.get('引流店铺页成交订单数'))),
            'store_gmv':        parse_num(r.get('引流店铺页用户支付金额(元)')),
            'sku_tag':          sku_tag,
            'hook_type':        hook_type,
            'content_pattern':  content_pattern,
            'composite_score':  0.0,  # 后续计算
            'annotation_source':'AUTO',
        }
        records.append(rec)
    return records


def calc_composite_scores(records: list[dict]) -> None:
    """归一化后计算composite_score"""
    keys = ['live_gmv', 'live_traffic', 'completion_rate', 'avg_watch_sec', 'post_search_gmv']
    max_vals = {}
    for k in keys:
        vals = [r[k] for r in records if r[k] is not None]
        max_vals[k] = max(vals) if vals else 1
    for r in records:
        r['composite_score'] = compute_composite_score(r, max_vals)


def generate_sql(records: list[dict], output_path: str) -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    cols = [
        'video_id', 'account_id', 'account_name', 'title', 'publish_time',
        'video_url', 'is_promoted', 'views', 'likes', 'comments', 'saves',
        'shares', 'follows', 'completion_rate', 'avg_watch_sec',
        'live_gmv', 'live_exposure', 'live_traffic', 'live_orders', 'live_cpm_gmv',
        'post_search_gmv', 'post_search_exp', 'post_search_click', 'post_search_orders',
        'direct_gmv', 'direct_orders', 'refund_amount', 'store_orders', 'store_gmv',
        'sku_tag', 'hook_type', 'content_pattern', 'composite_score', 'annotation_source',
        'create_at', 'create_by', 'create_name',
    ]

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("-- video_performance_record 历史数据导入\n")
        f.write(f"-- 生成时间: {now}\n")
        f.write(f"-- 记录数: {len(records)}\n")
        f.write("-- 执行前请确保已运行 V20260424__content_flywheel_foundation.sql\n\n")
        f.write("SET NAMES utf8mb4;\n")
        f.write("SET FOREIGN_KEY_CHECKS=0;\n\n")

        batch_size = 200
        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            f.write(f"INSERT INTO video_performance_record\n")
            f.write(f"  ({', '.join(cols)})\nVALUES\n")

            value_rows = []
            for r in batch:
                vals = []
                for c in cols:
                    if c == 'create_at':
                        vals.append(f"'{now}'")
                    elif c == 'create_by':
                        vals.append("0")
                    elif c == 'create_name':
                        vals.append("'import_script'")
                    elif c == 'is_promoted':
                        vals.append(str(r.get(c, 0)))
                    elif c in ('views', 'likes', 'comments', 'saves', 'shares', 'follows',
                               'avg_watch_sec', 'live_exposure', 'live_traffic', 'live_orders',
                               'post_search_exp', 'post_search_click', 'post_search_orders',
                               'direct_orders', 'store_orders'):
                        vals.append(str(int(r.get(c, 0) or 0)))
                    elif c in ('completion_rate', 'live_gmv', 'live_cpm_gmv', 'post_search_gmv',
                               'direct_gmv', 'refund_amount', 'store_gmv', 'composite_score'):
                        vals.append(str(round(float(r.get(c, 0) or 0), 6)))
                    elif c == 'publish_time':
                        v = r.get(c)
                        vals.append(f"'{v}'" if v else "NULL")
                    else:
                        vals.append(escape_sql(r.get(c)))
                value_rows.append(f"  ({', '.join(vals)})")

            f.write(",\n".join(value_rows))
            f.write(";\n\n")

        f.write("SET FOREIGN_KEY_CHECKS=1;\n")

    print(f"✅ SQL文件已生成: {output_path}")
    print(f"   共 {len(records)} 条INSERT记录，每批{batch_size}条")


def insert_db(records: list[dict], host, port, user, password, db) -> None:
    try:
        import pymysql
    except ImportError:
        print("错误: 需要 pymysql。执行: pip3 install pymysql")
        sys.exit(1)

    conn = pymysql.connect(host=host, port=int(port), user=user,
                           password=password, database=db,
                           charset='utf8mb4', autocommit=False)
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    cols = [
        'video_id', 'account_id', 'account_name', 'title', 'publish_time',
        'video_url', 'is_promoted', 'views', 'likes', 'comments', 'saves',
        'shares', 'follows', 'completion_rate', 'avg_watch_sec',
        'live_gmv', 'live_exposure', 'live_traffic', 'live_orders', 'live_cpm_gmv',
        'post_search_gmv', 'post_search_exp', 'post_search_click', 'post_search_orders',
        'direct_gmv', 'direct_orders', 'refund_amount', 'store_orders', 'store_gmv',
        'sku_tag', 'hook_type', 'content_pattern', 'composite_score', 'annotation_source',
        'create_at', 'create_by', 'create_name',
    ]

    placeholders = ', '.join(['%s'] * len(cols))
    sql = f"INSERT INTO video_performance_record ({', '.join(cols)}) VALUES ({placeholders})"

    batch = []
    for r in records:
        row = []
        for c in cols:
            if c == 'create_at':
                row.append(now)
            elif c == 'create_by':
                row.append(0)
            elif c == 'create_name':
                row.append('import_script')
            else:
                row.append(r.get(c))
        batch.append(tuple(row))

    try:
        cursor.executemany(sql, batch)
        conn.commit()
        print(f"✅ 已写入 {len(batch)} 条记录到数据库")
    except Exception as e:
        conn.rollback()
        print(f"❌ 写入失败: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def print_stats(records: list[dict]) -> None:
    from collections import Counter
    print("\n📊 标注统计:")
    sku_cnt = Counter(r['sku_tag'] for r in records)
    print("  SKU分布:")
    for k, v in sku_cnt.most_common():
        print(f"    {k}: {v}条")
    hook_cnt = Counter(r['hook_type'] for r in records)
    print("  Hook类型分布:")
    for k, v in hook_cnt.most_common():
        print(f"    {k}: {v}条")
    scores = sorted(r['composite_score'] for r in records)
    n = len(scores)
    if n:
        print(f"  composite_score: min={scores[0]:.4f} median={scores[n//2]:.4f} max={scores[-1]:.4f}")
    top5 = sorted(records, key=lambda r: r['composite_score'], reverse=True)[:5]
    print("  Top5 composite_score:")
    for r in top5:
        print(f"    [{r['hook_type']}|{r['sku_tag']}] score={r['composite_score']:.4f} live_gmv={r['live_gmv']:.0f} title={r['title'][:60]}")


def main():
    parser = argparse.ArgumentParser(description='视频历史数据导入脚本')
    parser.add_argument('--input', '-i', required=True, help='Excel文件路径')
    parser.add_argument('--mode', choices=['sql', 'db'], default='sql',
                        help='sql=生成SQL文件(默认) db=直连MySQL')
    parser.add_argument('--output', '-o',
                        default='scripts/output/video_performance_insert.sql',
                        help='SQL输出文件路径（mode=sql时有效）')
    parser.add_argument('--host', default='127.0.0.1')
    parser.add_argument('--port', default='3306')
    parser.add_argument('--user', default='root')
    parser.add_argument('--password', default='')
    parser.add_argument('--db', default='marketing_person')
    args = parser.parse_args()

    rows = load_excel(args.input)
    records = transform(rows)
    calc_composite_scores(records)
    print_stats(records)

    if args.mode == 'sql':
        generate_sql(records, args.output)
        print(f"\n下一步: 在MySQL客户端执行以下命令:")
        print(f"  source {args.output}")
        print(f"  -- 或: mysql -u root -p marketing_person < {args.output}")
    else:
        insert_db(records, args.host, args.port, args.user, args.password, args.db)


if __name__ == '__main__':
    main()
